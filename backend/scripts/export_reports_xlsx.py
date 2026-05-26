#!/usr/bin/env python3
"""Export per-component analysis from output/reports/*.json to Excel (.xlsx)."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from mr_framework.samples import PROJECT_ROOT

REPORTS_DIR = PROJECT_ROOT / "output" / "reports"
DEFAULT_OUT = PROJECT_ROOT / "output" / "analysis" / "component_analysis.xlsx"

TYPE_CATEGORIES = [
    "input_prop_relations",
    "state_event_relations",
    "interaction_accessibility_relations",
    "visual_layout_relations",
    "composition_context_relations",
    "data_flow_relations",
]

TC_LABELS = {
    "input_prop_relations": "输入属性",
    "state_event_relations": "状态事件",
    "interaction_accessibility_relations": "交互无障碍",
    "visual_layout_relations": "视觉布局",
    "composition_context_relations": "组合上下文",
    "data_flow_relations": "数据流",
}


def _load_reports() -> list[dict]:
    rows = []
    for path in sorted(REPORTS_DIR.glob("*.json")):
        try:
            rows.append(json.loads(path.read_text(encoding="utf-8")))
        except (json.JSONDecodeError, OSError):
            continue
    rows.sort(key=lambda r: int((r.get("meta") or {}).get("numeric_id") or 0))
    return rows


def _join_types(items: list | None, key: str = "relation_type") -> str:
    if not items:
        return ""
    parts = []
    for x in items:
        if isinstance(x, dict):
            parts.append(str(x.get(key, "")))
        else:
            parts.append(str(x))
    return "; ".join(p for p in parts if p)


def _component_row(r: dict) -> dict:
    m = r.get("meta") or {}
    c = r.get("completeness") or {}
    b = r.get("blind_spots") or {}
    by_tc = c.get("by_type_category") or {}
    priorities = r.get("test_priorities") or []
    suggestions = r.get("improvement_suggestions") or []

    row = {
        "序号": m.get("numeric_id"),
        "样本ID": m.get("sample_id"),
        "组件库": m.get("library"),
        "UI类别": m.get("category"),
        "组件名": m.get("component_name"),
        "MR总数": c.get("total_mrs"),
        "触达MR数": c.get("touched_mrs"),
        "严格覆盖MR数": c.get("covered_mrs"),
        "触达率": c.get("touch_rate"),
        "严格覆盖率": c.get("coverage_rate"),
        "加权触达率": c.get("weighted_touch_rate"),
        "加权严格覆盖率": c.get("weighted_coverage_rate"),
        "触达未覆盖数": c.get("touched_not_covered"),
        "触达未覆盖率": c.get("touched_not_covered_rate"),
        "完全未触达数": c.get("miss_at_uncov"),
        "未触达率": c.get("miss_at_uncov_rate"),
        "严格缺口数": c.get("strict_gap"),
        "严格缺口率": c.get("strict_gap_rate"),
        "关系类型数": c.get("distinct_relation_types"),
        "已覆盖关系类型数": c.get("covered_relation_types"),
        "未覆盖关系类型": _join_types(c.get("uncovered_relation_types")),
        "Open Bug压力": r.get("issue_pressure"),
        "盲区关系类型Top": _join_types(b.get("top_blind_mr_types")),
        "补测优先级条数": len(priorities),
        "改进建议条数": len(suggestions),
        "最高优先级类型": priorities[0].get("relation_type") if priorities else "",
        "最高优先级分数": priorities[0].get("priority_score") if priorities else "",
        "报告生成时间": r.get("generated_at"),
    }
    for tc in TYPE_CATEGORIES:
        label = TC_LABELS.get(tc, tc)
        block = by_tc.get(tc) or {}
        row[f"{label}_MR数"] = block.get("total")
        row[f"{label}_触达率"] = block.get("touch_rate")
        row[f"{label}_覆盖率"] = block.get("coverage_rate")
    return row


def _mr_rows(r: dict) -> list[dict]:
    m = r.get("meta") or {}
    cov_by_id = {x.get("mr_id"): x for x in (r.get("mr_coverage") or []) if x.get("mr_id")}
    out = []
    for mr in r.get("metamorphic_relations") or []:
        mr_id = mr.get("id")
        cov = cov_by_id.get(mr_id, {})
        out.append({
            "序号": m.get("numeric_id"),
            "样本ID": m.get("sample_id"),
            "组件库": m.get("library"),
            "UI类别": m.get("category"),
            "组件名": m.get("component_name"),
            "MR_ID": mr_id,
            "关系类型": mr.get("relation_type"),
            "大类": mr.get("type_category"),
            "置信度": mr.get("confidence"),
            "是否触达": cov.get("hybrid_touched", cov.get("touched")),
            "是否严格覆盖": cov.get("strict_covered", cov.get("covered")),
            "MR描述": (mr.get("description") or "")[:500],
            "期望关系": (mr.get("expected_relation") or "")[:300],
        })
    return out


def _blind_rows(r: dict) -> list[dict]:
    m = r.get("meta") or {}
    out = []
    for item in (r.get("blind_spots") or {}).get("blind_items") or []:
        out.append({
            "序号": m.get("numeric_id"),
            "样本ID": m.get("sample_id"),
            "组件库": m.get("library"),
            "UI类别": m.get("category"),
            "组件名": m.get("component_name"),
            "关系类型": item.get("relation_type"),
            "大类": item.get("type_category"),
            "严重程度": item.get("severity"),
            "原因": (item.get("reason") or "")[:400],
        })
    return out


def _priority_rows(r: dict) -> list[dict]:
    m = r.get("meta") or {}
    out = []
    for p in r.get("test_priorities") or []:
        out.append({
            "序号": m.get("numeric_id"),
            "样本ID": m.get("sample_id"),
            "组件库": m.get("library"),
            "组件名": m.get("component_name"),
            "关系类型": p.get("relation_type"),
            "大类": p.get("type_category"),
            "原因": p.get("reason"),
            "优先级分数": p.get("priority_score"),
            "测试变形": (p.get("test_transformation") or "")[:300],
            "期望Oracle": (p.get("expected_oracle") or "")[:300],
        })
    return out


def _write_sheet(ws, rows: list[dict], percent_cols: set[str]) -> None:
    if not rows:
        ws.append(["(无数据)"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    # Format percent columns
    try:
        from openpyxl.styles import numbers
        from openpyxl.utils import get_column_letter

        pct_fmt = numbers.FORMAT_PERCENTAGE_00
        for col_idx, h in enumerate(headers, start=1):
            if h in percent_cols or h.endswith("率") and h not in ("Open Bug压力",):
                if h.endswith("率") and "MR" not in h:
                    letter = get_column_letter(col_idx)
                    for row_idx in range(2, len(rows) + 2):
                        cell = ws[f"{letter}{row_idx}"]
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            cell.number_format = pct_fmt
            # Auto width (approximate)
            max_len = max(len(str(h)), *(len(str(row.get(h, ""))) for row in rows[:50]))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)
    except Exception:
        pass


def export_xlsx(out_path: Path) -> int:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    reports = _load_reports()
    if not reports:
        print("No reports in output/reports/", file=sys.stderr)
        return 1

    comp_rows = [_component_row(r) for r in reports]
    mr_rows: list[dict] = []
    blind_rows: list[dict] = []
    pri_rows: list[dict] = []
    for r in reports:
        mr_rows.extend(_mr_rows(r))
        blind_rows.extend(_blind_rows(r))
        pri_rows.extend(_priority_rows(r))

    rate_cols = {
        "触达率", "严格覆盖率", "加权触达率", "加权严格覆盖率",
        "触达未覆盖率", "未触达率", "严格缺口率",
    }
    for tc in TYPE_CATEGORIES:
        label = TC_LABELS.get(tc, tc)
        rate_cols.add(f"{label}_触达率")
        rate_cols.add(f"{label}_覆盖率")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "组件汇总"
    _write_sheet(ws1, comp_rows, rate_cols)

    ws2 = wb.create_sheet("MR明细")
    _write_sheet(ws2, mr_rows, set())

    ws3 = wb.create_sheet("盲区明细")
    _write_sheet(ws3, blind_rows, set())

    ws4 = wb.create_sheet("补测优先级")
    _write_sheet(ws4, pri_rows, set())

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  组件汇总: {len(comp_rows)} 行")
    print(f"  MR明细: {len(mr_rows)} 行")
    print(f"  盲区明细: {len(blind_rows)} 行")
    print(f"  补测优先级: {len(pri_rows)} 行")
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description="Export reports to Excel")
    p.add_argument("-o", "--output", type=Path, default=DEFAULT_OUT)
    args = p.parse_args()
    return export_xlsx(args.output)


if __name__ == "__main__":
    raise SystemExit(main())
