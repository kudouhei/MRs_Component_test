#!/usr/bin/env python3
"""Export issue-alignment statistical artifacts (charts + xlsx)."""

from __future__ import annotations

import argparse
import csv
import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
ALIGNMENT_JSON = PROJECT_ROOT / "output" / "aggregate" / "alignment_val.json"
OUT_DIR = PROJECT_ROOT / "output" / "analysis"
OUT_XLSX = OUT_DIR / "issue_alignment_stats.xlsx"
OUT_HTML = OUT_DIR / "issue_alignment_charts.html"


def _pct(v: float | None) -> str:
    if v is None:
        return "NA"
    return f"{100 * v:.2f}%"


def _f(v: float | None, n: int = 6) -> str:
    if v is None:
        return "NA"
    return f"{v:.{n}f}"


def _set_col_width(ws) -> None:
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in col[:80]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)


def _write_table(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    for row in rows:
        ws.append(row)
    _set_col_width(ws)


def _bar_svg(title: str, labels: list[str], values: list[float], *, max_val: float | None = None) -> str:
    width = 760
    row_h = 26
    top = 36
    left = 210
    h = top + len(labels) * row_h + 24
    right = 24
    chart_w = width - left - right
    mv = max_val if max_val is not None else (max(values) if values else 1.0)
    mv = max(mv, 1e-9)
    s = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{h}">',
        f'<text x="12" y="22" font-size="16" font-weight="600">{title}</text>',
    ]
    for i, (lab, val) in enumerate(zip(labels, values)):
        y = top + i * row_h
        bw = int(chart_w * (val / mv))
        s.append(f'<text x="10" y="{y + 16}" font-size="12">{lab}</text>')
        s.append(f'<rect x="{left}" y="{y + 3}" width="{bw}" height="14" fill="#3B82F6" rx="3" />')
        s.append(f'<text x="{left + bw + 8}" y="{y + 15}" font-size="12">{val:.3f}</text>')
    s.append("</svg>")
    return "\n".join(s)


def _build_html(payload: dict) -> str:
    exp1 = payload["exp1"]
    exp2 = payload["exp2"]
    bins = exp1.get("pressure_bins", [])
    high = exp1.get("high_pressure_vs_others_strict_gap_test", {})
    sig = exp2.get("topic_blind_significance", [])
    top_sig = sorted(sig, key=lambda x: (x.get("q_value", 1.0), x.get("p_value", 1.0)))[:10]

    chart_bins_gap = _bar_svg(
        "Mean strict_gap_rate by pressure bin",
        [f'{x["bin"]} (n={x["n"]})' for x in bins],
        [float(x.get("mean_strict_gap_rate", 0)) for x in bins],
        max_val=1.0,
    )
    chart_bins_cov = _bar_svg(
        "Mean coverage_rate by pressure bin",
        [f'{x["bin"]} (n={x["n"]})' for x in bins],
        [float(x.get("mean_coverage_rate", 0)) for x in bins],
        max_val=1.0,
    )
    chart_sig = _bar_svg(
        "Topic enrichment ratio (lower q first)",
        [f'{x["mr_type"]} (q={x.get("q_value", 1.0):.3f})' for x in top_sig],
        [float(x.get("enrichment_ratio") or 0.0) for x in top_sig],
    )

    rows_sig = []
    for x in top_sig:
        rows_sig.append(
            "<tr>"
            f"<td>{x['mr_type']}</td>"
            f"<td>{x['overlap_count']}</td>"
            f"<td>{_f(x.get('expected_overlap_under_null'), 3)}</td>"
            f"<td>{_f(x.get('enrichment_ratio'), 3)}</td>"
            f"<td>{_f(x.get('odds_ratio'), 3)}</td>"
            f"<td>{_f(x.get('p_value'), 6)}</td>"
            f"<td>{_f(x.get('q_value'), 6)}</td>"
            "</tr>"
        )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Issue Alignment Statistical Evidence</title>
  <style>
    body {{ font-family: -apple-system,BlinkMacSystemFont,Segoe UI,Arial,sans-serif; margin: 24px; max-width: 980px; }}
    h1 {{ font-size: 24px; }} h2 {{ margin-top: 24px; }}
    .cards {{ display: grid; grid-template-columns: repeat(4, minmax(120px,1fr)); gap: 12px; }}
    .card {{ border: 1px solid #E5E7EB; border-radius: 8px; padding: 12px; background: #FAFAFA; }}
    .card .v {{ font-size: 20px; font-weight: 700; }}
    .card .k {{ color: #4B5563; font-size: 12px; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 10px; }}
    th, td {{ border: 1px solid #E5E7EB; padding: 6px 10px; font-size: 13px; }}
    th {{ background: #F3F4F6; text-align: left; }}
    .muted {{ color: #6B7280; }}
  </style>
</head>
<body>
  <h1>Issue Alignment: Statistical Evidence</h1>
  <p class="muted">Source: output/aggregate/alignment_val.json</p>

  <div class="cards">
    <div class="card"><div class="v">{exp1.get('n')}</div><div class="k">Samples</div></div>
    <div class="card"><div class="v">{high.get('n_high', 0)}</div><div class="k">High Pressure (>=10)</div></div>
    <div class="card"><div class="v">{_f(high.get('strict_gap_diff_high_minus_other'), 3)}</div><div class="k">Gap Diff (high-other)</div></div>
    <div class="card"><div class="v">{_f(high.get('strict_gap_diff_permutation_pvalue'), 4)}</div><div class="k">Permutation p-value</div></div>
  </div>

  <h2>Pressure Buckets</h2>
  {chart_bins_gap}
  {chart_bins_cov}

  <h2>High Pressure vs Others Test</h2>
  <table>
    <tr><th>Metric</th><th>Value</th></tr>
    <tr><td>mean_strict_gap_high</td><td>{_f(high.get('mean_strict_gap_high'), 6)}</td></tr>
    <tr><td>mean_strict_gap_other</td><td>{_f(high.get('mean_strict_gap_other'), 6)}</td></tr>
    <tr><td>strict_gap_diff_high_minus_other</td><td>{_f(high.get('strict_gap_diff_high_minus_other'), 6)}</td></tr>
    <tr><td>bootstrap95ci</td><td>{high.get('strict_gap_diff_bootstrap95ci')}</td></tr>
    <tr><td>permutation_pvalue</td><td>{_f(high.get('strict_gap_diff_permutation_pvalue'), 6)}</td></tr>
    <tr><td>cliffs_delta</td><td>{_f(high.get('strict_gap_cliffs_delta'), 6)}</td></tr>
  </table>

  <h2>Topic Alignment Significance</h2>
  {chart_sig}
  <table>
    <tr>
      <th>mr_type</th><th>overlap</th><th>expected</th><th>enrichment</th><th>odds_ratio</th><th>p</th><th>q(FDR)</th>
    </tr>
    {''.join(rows_sig)}
  </table>
</body>
</html>"""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=ALIGNMENT_JSON)
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"Missing alignment file: {args.input}")

    payload = json.loads(args.input.read_text(encoding="utf-8"))
    exp1 = payload.get("exp1", {})
    exp2 = payload.get("exp2", {})

    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    # CSV exports
    bins = exp1.get("pressure_bins", [])
    with (out_dir / "issue_pressure_bins.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["bin", "n", "mean_coverage_rate", "mean_strict_gap_rate", "zero_coverage_count"],
        )
        w.writeheader()
        for row in bins:
            w.writerow(row)

    sig = exp2.get("topic_blind_significance", [])
    with (out_dir / "issue_topic_significance.csv").open("w", encoding="utf-8", newline="") as f:
        fieldnames = [
            "mr_type",
            "n_pairs",
            "topic_count",
            "blind_count",
            "overlap_count",
            "expected_overlap_under_null",
            "enrichment_ratio",
            "odds_ratio",
            "p_value",
            "q_value",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in sig:
            w.writerow({k: row.get(k) for k in fieldnames})

    # XLSX export
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "overview"
    high = exp1.get("high_pressure_vs_others_strict_gap_test", {})
    _write_table(
        ws0,
        ["metric", "value"],
        [
            ["n_samples", exp1.get("n")],
            ["pearson_pressure_vs_coverage", exp1.get("pearson_pressure_vs_coverage_rate")],
            ["pearson_pressure_vs_strict_gap", exp1.get("pearson_pressure_vs_strict_gap_rate")],
            ["spearman_pressure_vs_coverage", exp1.get("spearman_pressure_vs_coverage_rate")],
            ["spearman_pressure_vs_strict_gap", exp1.get("spearman_pressure_vs_strict_gap_rate")],
            ["high_threshold", high.get("high_threshold")],
            ["n_high", high.get("n_high")],
            ["n_other", high.get("n_other")],
            ["mean_strict_gap_high", high.get("mean_strict_gap_high")],
            ["mean_strict_gap_other", high.get("mean_strict_gap_other")],
            ["strict_gap_diff_high_minus_other", high.get("strict_gap_diff_high_minus_other")],
            ["strict_gap_diff_permutation_pvalue", high.get("strict_gap_diff_permutation_pvalue")],
            ["strict_gap_diff_bootstrap95ci", str(high.get("strict_gap_diff_bootstrap95ci"))],
            ["strict_gap_cliffs_delta", high.get("strict_gap_cliffs_delta")],
            ["n_matched_issue_component_pairs", exp2.get("n_matched_issue_component_pairs")],
            ["significant_topics_fdr_005", "; ".join(exp2.get("significant_topics_fdr_005") or [])],
        ],
    )

    ws1 = wb.create_sheet("pressure_bins")
    _write_table(
        ws1,
        ["bin", "n", "mean_coverage_rate", "mean_strict_gap_rate", "zero_coverage_count"],
        [
            [
                row.get("bin"),
                row.get("n"),
                row.get("mean_coverage_rate"),
                row.get("mean_strict_gap_rate"),
                row.get("zero_coverage_count"),
            ]
            for row in bins
        ],
    )

    ws2 = wb.create_sheet("high_vs_others_test")
    _write_table(
        ws2,
        ["metric", "value"],
        [[k, (str(v) if isinstance(v, (list, dict)) else v)] for k, v in high.items()],
    )

    ws3 = wb.create_sheet("topic_significance")
    _write_table(
        ws3,
        [
            "mr_type",
            "n_pairs",
            "topic_count",
            "blind_count",
            "overlap_count",
            "expected_overlap_under_null",
            "enrichment_ratio",
            "odds_ratio",
            "p_value",
            "q_value",
        ],
        [
            [
                row.get("mr_type"),
                row.get("n_pairs"),
                row.get("topic_count"),
                row.get("blind_count"),
                row.get("overlap_count"),
                row.get("expected_overlap_under_null"),
                row.get("enrichment_ratio"),
                row.get("odds_ratio"),
                row.get("p_value"),
                row.get("q_value"),
            ]
            for row in sig
        ],
    )

    ws4 = wb.create_sheet("top_topic_overlaps")
    overlaps = exp2.get("top_topic_blind_overlaps", [])
    _write_table(
        ws4,
        ["mr_type", "overlap_count", "issue_topic_count", "overlap_ratio"],
        [
            [row.get("mr_type"), row.get("overlap_count"), row.get("issue_topic_count"), row.get("overlap_ratio")]
            for row in overlaps
        ],
    )

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    # HTML charts
    OUT_HTML.write_text(_build_html(payload), encoding="utf-8")

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_HTML}")
    print(f"Wrote {out_dir / 'issue_pressure_bins.csv'}")
    print(f"Wrote {out_dir / 'issue_topic_significance.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

